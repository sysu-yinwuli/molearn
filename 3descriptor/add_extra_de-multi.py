# ===================== CONFIG 区域（写完整文件名） =====================
EXCEL_FILES = [                     # 10 个分表，写完整
    'xtb_data_bsi1-zr-extracted.xlsx',
    'xtb_data_bsi1-ti-extracted.xlsx',
    'xtb_data_bc1-zr-extracted.xlsx',
    'xtb_data_bc1-ti-extracted.xlsx',
    'xtb_data_b3-extracted.xlsx',
    'xtb_data_b2-zr-extracted.xlsx',
    'xtb_data_b2-ti-extracted.xlsx',
    'xtb_data_b1-hf-extracted.xlsx',
    'xtb_data_b2-hf-extracted.xlsx',
    'xtb_data_b0-extracted.xlsx'
]
IN_NPY      = 'poly-all.npy'         # 原始 44 W 分子
OUT_NPY     = 'poly-all-qc.npy'  # 输出文件
DE_COLS     = list(range(1, 34))    # Excel 第 1–33 列（索引从 0 算）
# ===================================================================

import openpyxl
import numpy as np

# ---------------- 1. 按完整文件名依次读取 ----------------
m_name_y = {}      # 分子名 -> 33 维描述符
header   = None    # 列名只拿第一份表

for excel in EXCEL_FILES:
    wb = openpyxl.load_workbook(excel)
    ws = wb.active
    # 拿表头（仅第一次）
    if header is None:
        header = [ws[1][d].value for d in DE_COLS]
    # 遍历数据行
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if key is None:
            continue
        key = str(key).split('.')[0]          # 去掉扩展名
        m_name_y[key] = [row[d] for d in DE_COLS]
    wb.close()
print(f'Excel 读取完成，共 {len(m_name_y)} 条描述符')

# ---------------- 2. 加载 .npy 数据 ----------------
data = np.load(IN_NPY, allow_pickle=True)
if data.ndim == 0:            # 处理 0-d 数组
    data = data.item()
# 兼容 {'successful':[...]} 或纯列表
if isinstance(data, dict) and 'successful' in data:
    molecules = data['successful']
else:
    molecules = data
print(f'NPY 加载完成，共 {len(molecules)} 个分子')

# ---------------- 3. 注入描述符 ----------------
for d in molecules:
    key = d['name'].split('.')[0]
    if key in m_name_y:
        d['extra_d']     = m_name_y[key]
        d['name_of_extra'] = header
    else:
        print(f'[Warning] {key} 缺失描述符，已跳过')

# ---------------- 4. 保存新 .npy ----------------
to_save = data if isinstance(data, dict) else molecules
np.save(OUT_NPY, to_save, allow_pickle=True)
print(f'全部完成！已保存为 ./{OUT_NPY}')